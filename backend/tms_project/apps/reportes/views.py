"""
apps/reportes/views.py — Generación y exportación de reportes
"""
import io
import csv
import logging
from datetime import date
from django.http import HttpResponse, FileResponse
from django.db.models import Sum, Count
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tms_project.apps.ingresos.models import Ingreso, Gasto
from tms_project.apps.viajes.models import Viaje, EstadoViaje

logger = logging.getLogger("tms")

COLOR_HEADER = colors.HexColor("#1a5276")
COLOR_ALT = colors.HexColor("#d6eaf8")


def _formatear_pyg(valor):
    """Formatea un número como Guaraní paraguayo."""
    if valor is None:
        return "₲ 0"
    return f"₲ {int(valor):,}".replace(",", ".")


class ReporteViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def _parse_fechas(self, request):
        desde = request.query_params.get("desde", date.today().replace(day=1).isoformat())
        hasta = request.query_params.get("hasta", date.today().isoformat())
        return date.fromisoformat(desde), date.fromisoformat(hasta)

    def _datos_resumen(self, desde, hasta):
        ingresos = Ingreso.objects.filter(fecha__range=(desde, hasta))
        gastos = Gasto.objects.filter(fecha__range=(desde, hasta))
        viajes = Viaje.objects.filter(
            fecha_salida__range=(desde, hasta),
            estado=EstadoViaje.FINALIZADO,
        )
        total_ing = ingresos.aggregate(t=Sum("monto"))["t"] or 0
        total_gas = gastos.aggregate(t=Sum("monto"))["t"] or 0
        return {
            "ingresos": ingresos,
            "gastos": gastos,
            "viajes": viajes,
            "total_ingresos": total_ing,
            "total_gastos": total_gas,
            "ganancia_neta": total_ing - total_gas,
        }

    # ========================
    # PDF
    # ========================
    @action(detail=False, methods=["get"], url_path="pdf")
    def exportar_pdf(self, request):
        desde, hasta = self._parse_fechas(request)
        tipo = request.query_params.get("tipo", "resumen")
        datos = self._datos_resumen(desde, hasta)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1.5*cm, leftMargin=1.5*cm,
            topMargin=2*cm, bottomMargin=1.5*cm,
        )
        styles = getSampleStyleSheet()
        titulo_style = ParagraphStyle(
            "Titulo", parent=styles["Title"],
            fontSize=16, textColor=COLOR_HEADER, spaceAfter=8,
        )
        subtitulo_style = ParagraphStyle(
            "Subtitulo", parent=styles["Normal"],
            fontSize=10, textColor=colors.grey, spaceAfter=16,
        )

        elements = []
        elements.append(Paragraph("Sistema TMS — Reporte de Gestión de Transporte", titulo_style))
        elements.append(Paragraph(
            f"Período: {desde.strftime('%d/%m/%Y')} al {hasta.strftime('%d/%m/%Y')}",
            subtitulo_style,
        ))

        # Tabla resumen financiero
        resumen_data = [
            ["Indicador", "Valor"],
            ["Total Ingresos", _formatear_pyg(datos["total_ingresos"])],
            ["Total Gastos", _formatear_pyg(datos["total_gastos"])],
            ["Ganancia Neta", _formatear_pyg(datos["ganancia_neta"])],
            ["Viajes Realizados", str(datos["viajes"].count())],
            ["Margen (%)", f"{float(datos['ganancia_neta']) / float(datos['total_ingresos']) * 100:.1f}%" if datos["total_ingresos"] else "0%"],
        ]
        tabla_resumen = Table(resumen_data, colWidths=[8*cm, 6*cm])
        tabla_resumen.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), COLOR_HEADER),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, COLOR_ALT]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(tabla_resumen)
        elements.append(Spacer(1, 0.5*cm))

        # Tabla de ingresos
        if datos["ingresos"].exists():
            elements.append(Paragraph("Detalle de Ingresos", styles["Heading2"]))
            ing_data = [["Fecha", "Cliente", "Viaje", "Forma de Pago", "Monto"]]
            for ing in datos["ingresos"].select_related("cliente", "viaje").order_by("-fecha")[:50]:
                ing_data.append([
                    ing.fecha.strftime("%d/%m/%Y"),
                    ing.cliente.razon_social[:30],
                    ing.viaje.numero_viaje if ing.viaje else "—",
                    ing.get_forma_pago_display(),
                    _formatear_pyg(ing.monto),
                ])
            t = Table(ing_data, colWidths=[2.5*cm, 7*cm, 3*cm, 4*cm, 4*cm])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), COLOR_HEADER),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, COLOR_ALT]),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
                ("ALIGN", (4, 0), (4, -1), "RIGHT"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]))
            elements.append(t)

        doc.build(elements)
        buffer.seek(0)
        return FileResponse(
            buffer,
            as_attachment=True,
            filename=f"reporte_tms_{desde}_{hasta}.pdf",
            content_type="application/pdf",
        )

    # ========================
    # EXCEL
    # ========================
    @action(detail=False, methods=["get"], url_path="excel")
    def exportar_excel(self, request):
        desde, hasta = self._parse_fechas(request)
        datos = self._datos_resumen(desde, hasta)

        wb = Workbook()
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="1A5276")
        alt_fill = PatternFill("solid", fgColor="D6EAF8")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center = Alignment(horizontal="center", vertical="center")
        right = Alignment(horizontal="right")

        # ——— Hoja Resumen ———
        ws = wb.active
        ws.title = "Resumen"
        ws.append(["SISTEMA TMS — REPORTE DE GESTIÓN"])
        ws.append([f"Período: {desde.strftime('%d/%m/%Y')} al {hasta.strftime('%d/%m/%Y')}"])
        ws.append([])
        ws.append(["Indicador", "Valor"])
        resumen_rows = [
            ("Total Ingresos (₲)", float(datos["total_ingresos"])),
            ("Total Gastos (₲)", float(datos["total_gastos"])),
            ("Ganancia Neta (₲)", float(datos["ganancia_neta"])),
            ("Viajes Realizados", datos["viajes"].count()),
        ]
        for row in resumen_rows:
            ws.append(list(row))

        for cell in ws[4]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # ——— Hoja Ingresos ———
        ws_ing = wb.create_sheet("Ingresos")
        headers = ["Fecha", "Cliente", "RUC", "Viaje", "Forma de Pago", "Factura", "Monto (₲)", "Observaciones"]
        ws_ing.append(headers)
        for i, h in enumerate(headers, 1):
            cell = ws_ing.cell(1, i)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for r, ing in enumerate(datos["ingresos"].select_related("cliente", "viaje").order_by("-fecha"), 2):
            ws_ing.append([
                ing.fecha.strftime("%d/%m/%Y"),
                ing.cliente.razon_social,
                ing.cliente.ruc,
                ing.viaje.numero_viaje if ing.viaje else "",
                ing.get_forma_pago_display(),
                ing.numero_factura,
                float(ing.monto),
                ing.observaciones,
            ])
            if r % 2 == 0:
                for c in range(1, len(headers) + 1):
                    ws_ing.cell(r, c).fill = alt_fill

        # ——— Hoja Gastos ———
        ws_gas = wb.create_sheet("Gastos")
        headers_g = ["Fecha", "Categoría", "Viaje", "Comprobante", "Proveedor", "Monto (₲)", "Descripción"]
        ws_gas.append(headers_g)
        for i, h in enumerate(headers_g, 1):
            cell = ws_gas.cell(1, i)
            cell.font = header_font
            cell.fill = header_fill

        for r, g in enumerate(datos["gastos"].select_related("viaje").order_by("-fecha"), 2):
            ws_gas.append([
                g.fecha.strftime("%d/%m/%Y"),
                g.get_categoria_display(),
                g.viaje.numero_viaje if g.viaje else "",
                g.numero_comprobante,
                g.proveedor,
                float(g.monto),
                g.descripcion,
            ])

        # Auto-ancho de columnas
        for ws_sheet in [ws, ws_ing, ws_gas]:
            for col in ws_sheet.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws_sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="reporte_tms_{desde}_{hasta}.xlsx"'},
        )

    # ========================
    # CSV
    # ========================
    @action(detail=False, methods=["get"], url_path="csv")
    def exportar_csv(self, request):
        desde, hasta = self._parse_fechas(request)
        tipo = request.query_params.get("tipo", "ingresos")

        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="tms_{tipo}_{desde}_{hasta}.csv"'
        response.write("\ufeff")  # BOM para Excel

        writer = csv.writer(response)
        if tipo == "ingresos":
            writer.writerow(["Fecha", "Cliente", "RUC", "Viaje", "Forma Pago", "Monto", "Moneda", "Factura"])
            for ing in Ingreso.objects.filter(fecha__range=(desde, hasta)).select_related("cliente", "viaje"):
                writer.writerow([
                    ing.fecha, ing.cliente.razon_social, ing.cliente.ruc,
                    ing.viaje.numero_viaje if ing.viaje else "",
                    ing.forma_pago, ing.monto, ing.moneda, ing.numero_factura,
                ])
        elif tipo == "gastos":
            writer.writerow(["Fecha", "Categoría", "Viaje", "Monto", "Moneda", "Comprobante", "Descripción"])
            for g in Gasto.objects.filter(fecha__range=(desde, hasta)).select_related("viaje"):
                writer.writerow([
                    g.fecha, g.get_categoria_display(),
                    g.viaje.numero_viaje if g.viaje else "",
                    g.monto, g.moneda, g.numero_comprobante, g.descripcion,
                ])
        elif tipo == "viajes":
            writer.writerow(["N° Viaje", "Fecha Salida", "Fecha Regreso", "Origen", "Destino", "Cliente", "Chofer", "Km", "Estado"])
            for v in Viaje.objects.filter(fecha_salida__range=(desde, hasta)).select_related("cliente", "chofer"):
                writer.writerow([
                    v.numero_viaje, v.fecha_salida, v.fecha_regreso or "",
                    v.origen, v.destino, v.cliente.razon_social,
                    v.chofer.nombre_completo, v.distancia_recorrida or "",
                    v.get_estado_display(),
                ])
        return response
